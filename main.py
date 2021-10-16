from openpyxl.workbook.workbook import Workbook

from openpyxl import load_workbook


def not_have(x, *values):
    if x is None:
        return True

    no = True
    for value in values:
        no = no and not (value in x)
    return no


def have(x, *values):
    if x is None:
        return True

    yes = False
    for value in values:
        yes = yes or value in x
    return yes


def must_none(x):
    return x is None


# 修改自己需要的内容

files = [
    {
        'case': '国考.xlsx',  # 查找文件路径
        'skip_rows': [1, 2],  # 需要跳过的行号
        'merge_row': 2,  # 存在 key 但被合并的行好
        'keys_row': 2,  # key 所在行号
        'exports': [
            {
                'file_name': "B-国考.xlsx",  # 导出文件路径
                'filters': {
                    # 以下就是筛选条件
                    '专业': lambda x: have(x, '机械', '工学', '不限'),
                    '学历': lambda x: have(x, '本科', '大专及以上'),
                    '学位': lambda x: x != '硕士',
                    '政治面貌': lambda x: have(x, '不限'),
                    '服务基层项目工作经历': lambda x: have(x, '无限制', '不限'),
                    '基层工作最低年限': lambda x: have(x, '无限制', '不限'),
                    '工作地点': lambda x: have(x, '广东'),
                    '落户地点': lambda x: have(x, '广东'),
                    '备注': lambda x: not_have(x, '女性', '至少具有注册会计师', '大学英语'),
                }
            },
        ]
    },
    {
        'case': '深圳事业单位.xlsx',
        'skip_rows': [1, 2, 3],
        'merge_row': 2,
        'keys_row': 3,
        'exports': [
            {
                'file_name': "B-深圳事业单位.xlsx",
                'filters': {
                    '专业': lambda x: have(x, '机械', '工学', '本科：不限'),
                    '最低专业技术资格': lambda x: must_none(x),
                    '学历': lambda x: have(x, '本科'),
                    '学位': lambda x: have(x, '学士'),
                    '与岗位有关的其它条件': lambda x: not_have(x, '女性', '中共党员', '证', '资格'),
                    '笔试类别': lambda x: not_have(x, '社会人员'),
                }
            },
        ]
    },
]


def add_rows(rows, sheet, merges=None):
    row_index = sheet.max_row + 1 if merges is None else sheet.max_row
    for cell_model in rows:
        if cell_model.value is None and not (merges is None) and cell_model.column < len(merges):
            sheet.cell(column=cell_model.column, row=row_index, value=merges[cell_model.column - 1].value)
        else:
            sheet.cell(column=cell_model.column, row=row_index, value=cell_model.value)


def reset_col(sheet):
    for col in sheet.columns:
        max_cell = max(col, default=0, key=lambda a: len(str(a.value).encode('gbk')))
        max_len = len(str(max_cell.value).encode('gbk')) + 2
        sheet.column_dimensions[col[0].column_letter].width = max_len


for file in files:
    case = file['case']
    exports = file['exports']
    skip_rows = file['skip_rows']
    keys_row = file['keys_row']
    merge_row = file['merge_row']
    source_excel = load_workbook(case)

    for export in exports:

        filters = export['filters']
        export_excel = Workbook()
        export_excel.remove(export_excel.active)

        for source_sheet in source_excel.worksheets:
            export_sheet = export_excel.create_sheet(source_sheet.title)
            index = 0
            for row in source_sheet.rows:
                index += 1
                if index in skip_rows:
                    if index == keys_row:
                        add_rows(row, export_sheet, list(source_sheet.rows)[merge_row - 1])
                    continue

                can_write = True
                for cell in row:
                    key = source_sheet.cell(keys_row, cell.col_idx)
                    if key.value is None:
                        key = source_sheet.cell(merge_row, cell.col_idx)
                    if key.value in filters and not filters[key.value](cell.value):
                        can_write = False
                        break

                if can_write:
                    add_rows(row, export_sheet)
            reset_col(export_sheet)
        file_name = export['file_name']
        export_excel.save(file_name)
